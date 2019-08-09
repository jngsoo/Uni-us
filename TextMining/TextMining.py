from konlpy.tag import Kkma
from openpyxl import load_workbook
import collections
import re

kkma = Kkma()

load_wb = load_workbook("../result.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

# content = []
# file = open('content.txt')
#
# time = 0
#
# for line in file:
#     if time==1000:
#         break
#     tagList = kkma.pos(line)
#
#     for token in tagList:
#         if time == 1000:
#             break
#         if token[1]=='NNG':
#             content.append(token[0])
#             time += 1
#
# nounDict = collections.Counter(content)
# nounDict = dict(nounDict)
# realDict = collections.Counter(nounDict)
# print(realDict)




wordDict = {}

def cleanText(rawText):
    pattern = re.compile('\w*\s*')
    result = "".join(re.findall(pattern, rawText))
    return result


# 셀 좌표로 값 출력
for i in range(2, 3000):
    titles = load_ws.cell(i, 3).value
    titles = cleanText(titles)

    contents = load_ws.cell(i, 4).value
    contents = cleanText(contents)

    texts = titles + contents
    malist = kkma.pos(texts)

    for x in malist:
        if 'NNG' in x:
            keyword = x[0]
            if keyword in wordDict:
                wordDict[keyword] += 1
            else:
                wordDict[keyword] = 1

wordDict = sorted(wordDict.items(), key=lambda t: t[1], reverse=True)

print(wordDict)